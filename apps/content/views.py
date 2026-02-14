from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse
import csv
import io


from apps.core.context_processors import navbar_context
from apps.core.permissions import is_content_manager_or_admin
from .models import Occupation, OccupationTask, Industry, Skill, OccupationMedia
from .forms import (
    OccupationForm,
    OccupationTaskForm,
    IndustryForm,
    SkillForm,
    OccupationMediaForm,
)


@login_required
@user_passes_test(is_content_manager_or_admin)
def occupation_edit(request, occupation_id):
    """
    Main view for editing an occupation.
    Renders the page shell which loads partials via HTMX?
    Strictly following the profile pattern, the main view loads the page with the first tab content pre-loaded or ready to load.
    Here we'll load the full page with the details form ready.
    """
    occupation = get_object_or_404(Occupation, pk=occupation_id)

    # We can just render the template. The template will include the
    # partials or use hx-trigger="load" to fetch them.
    # profile.html uses hx-trigger="load" for the tab content.

    context = navbar_context(request)
    context["occupation"] = occupation
    return render(request, "content/occupation_edit.html", context)


@login_required
@user_passes_test(is_content_manager_or_admin)
def occupation_details_partial(request, occupation_id):
    """
    HTMX partial for editing occupation details.
    """
    occupation = get_object_or_404(Occupation, pk=occupation_id)

    if request.method == "POST":
        form = OccupationForm(request.POST, instance=occupation)
        if form.is_valid():
            form.save()
            messages.success(request, "Occupation details updated successfully!")
    else:
        form = OccupationForm(instance=occupation)

    context = {"form": form, "occupation": occupation}
    return render(request, "content/partials/occupation_details.html", context)


@login_required
@user_passes_test(is_content_manager_or_admin)
def occupation_tasks_partial(request, occupation_id):
    """
    HTMX partial for listing and adding tasks.
    """
    occupation = get_object_or_404(Occupation, pk=occupation_id)

    if request.method == "POST":
        form = OccupationTaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.occupation = occupation
            task.save()
            messages.success(request, "Task created successfully!")
            form = OccupationTaskForm()  # Reset form
    else:
        form = OccupationTaskForm()

    tasks = occupation.tasks.all()

    context = {"form": form, "occupation": occupation, "tasks": tasks}
    return render(request, "content/partials/occupation_tasks.html", context)


@login_required
@user_passes_test(is_content_manager_or_admin)
def occupation_media_partial(request, occupation_id):
    """
    HTMX partial for listing and adding occupation media.
    """
    occupation = get_object_or_404(Occupation, pk=occupation_id)

    if request.method == "POST":
        form = OccupationMediaForm(request.POST)
        if form.is_valid():
            media = form.save(commit=False)
            media.occupation = occupation
            media.save()
            messages.success(request, "Media item added successfully!")
            form = OccupationMediaForm()
    else:
        form = OccupationMediaForm()

    media_items = OccupationMedia.objects.filter(occupation=occupation)

    context = {
        "form": form,
        "occupation": occupation,
        "media_items": media_items,
    }
    return render(request, "content/partials/occupation_media.html", context)


@login_required
@user_passes_test(is_content_manager_or_admin)
def occupation_media_detail(request, occupation_id, media_id):
    """
    HTMX partial for editing/deleting a single media item.
    """
    occupation = get_object_or_404(Occupation, pk=occupation_id)
    media_item = get_object_or_404(OccupationMedia, pk=media_id, occupation=occupation)

    if request.method == "DELETE":
        media_item.delete()
        messages.success(request, "Media item deleted.")
        return HttpResponse()

    start_editing = request.GET.get("mode") == "edit"

    if request.method == "POST":
        form = OccupationMediaForm(request.POST, instance=media_item)
        if form.is_valid():
            form.save()
            messages.success(request, "Media item updated.")
            start_editing = False
    else:
        form = OccupationMediaForm(instance=media_item)

    context = {
        "media_item": media_item,
        "form": form if start_editing else None,
        "start_editing": start_editing,
        "occupation": occupation,
    }
    template_name = (
        "content/partials/item_media_edit.html"
        if start_editing
        else "content/partials/item_media.html"
    )
    return render(request, template_name, context)


@login_required
@user_passes_test(is_content_manager_or_admin)
def occupation_task_detail(request, occupation_id, task_id):
    """
    HTMX partial for editing/deleting a single task.
    """
    occupation = get_object_or_404(Occupation, pk=occupation_id)
    task = get_object_or_404(OccupationTask, pk=task_id, occupation=occupation)

    if request.method == "DELETE":
        task.delete()
        messages.success(request, "Task deleted.")
        return HttpResponse()  # Client side should remove the row

    # Editing (using GET with ?mode=edit or pure POST)
    # The profile example uses a query param ?mode=edit to switch to edit mode

    start_editing = request.GET.get("mode") == "edit"

    if request.method == "POST":
        form = OccupationTaskForm(request.POST, instance=task)
        if form.is_valid():
            form.save()
            messages.success(request, "Task updated.")
            start_editing = False  # Exit edit mode
    else:
        # GET request, possibly just viewing or fetching form for edit
        form = OccupationTaskForm(instance=task)

    context = {
        "task": task,
        "form": form if start_editing else None,  # Only pass form if editing
        "start_editing": start_editing,
        "occupation": occupation,
    }
    return render(request, "content/partials/item_task.html", context)


@login_required
@user_passes_test(is_content_manager_or_admin)
def occupation_add(request):
    """
    View to add a single occupation with tasks and potentially a new industry.
    """
    if request.method == "POST":
        form = OccupationForm(request.POST)

        # Handle dynamic industry name if it doesn't exist as an ID
        industry_input = request.POST.get("industry_name")
        industry_id = request.POST.get("industry")

        if not industry_id and industry_input:
            # Try to find industry by name or create it
            industry, created = Industry.objects.get_or_create(
                name=industry_input,
                defaults={"code": industry_input.upper().replace(" ", "_")[:50]},
            )
            # Link it to the form data
            data = request.POST.copy()
            data["industry"] = industry.id
            form = OccupationForm(data)

        if form.is_valid():
            occupation = form.save()

            # Handle tasks
            task_titles = request.POST.getlist("task_titles[]")
            task_descriptions = request.POST.getlist("task_descriptions[]")

            for title, desc in zip(task_titles, task_descriptions):
                if title.strip():
                    OccupationTask.objects.create(
                        occupation=occupation,
                        title=title.strip(),
                        description=desc.strip(),
                    )

            messages.success(request, "Occupation created successfully with tasks!")
            return redirect("occupation-edit", occupation_id=occupation.id)
    else:
        form = OccupationForm()

    # Get some industries for autocomplete
    industries = Industry.objects.all().values("id", "name")[:20]

    # Get some common tasks for autocomplete (title and description)
    common_tasks = list(
        OccupationTask.objects.values("title", "description").distinct()[:50]
    )

    import json

    context = navbar_context(request)
    context["form"] = form
    context["industries_json"] = json.dumps(list(industries))
    context["common_tasks_json"] = json.dumps(list(common_tasks))
    context["nqf_levels"] = [
        {"id": "0", "name": "Any / Not Applicable (0)"},
        {"id": "4", "name": "Grade 12 / Matric (4)"},
        {"id": "5", "name": "Higher Certificate (5)"},
        {"id": "6", "name": "Diploma / Advanced Certificate (6)"},
        {"id": "7", "name": "Bachelor's Degree / Advanced Diploma (7)"},
        {"id": "8", "name": "Honours Degree / PG Diploma (8)"},
        {"id": "9", "name": "Master's Degree (9)"},
        {"id": "10", "name": "Doctoral Degree (10)"},
    ]
    return render(request, "content/occupation_add.html", context)


@login_required
@user_passes_test(is_content_manager_or_admin)
def occupation_upload(request):
    """
    View to bulk upload occupations via Excel.
    """
    if request.GET.get("template") == "excel":
        from django.http import HttpResponse
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation

        industries = list(Industry.objects.values("code", "name").order_by("name"))
        nqf_levels = [
            ("0", "Any / Not Applicable (0)"),
            ("4", "Grade 12 / Matric (4)"),
            ("5", "Higher Certificate (5)"),
            ("6", "Diploma / Advanced Certificate (6)"),
            ("7", "Bachelor's Degree / Advanced Diploma (7)"),
            ("8", "Honours Degree / PG Diploma (8)"),
            ("9", "Master's Degree (9)"),
            ("10", "Doctoral Degree (10)"),
        ]

        workbook = Workbook()
        template_sheet = workbook.active
        template_sheet.title = "Template"
        template_sheet.append(
            [
                "ofo_code",
                "ofo_title",
                "description",
                "industry_code",
                "years_of_experience",
                "preferred_nqf_level",
            ]
        )

        industries_sheet = workbook.create_sheet("Industries")
        industries_sheet.append(["code", "name"])
        for industry in industries:
            industries_sheet.append([industry["code"], industry["name"]])

        nqf_sheet = workbook.create_sheet("NQF Levels")
        nqf_sheet.append(["level", "label"])
        for level, label in nqf_levels:
            nqf_sheet.append([level, label])

        if industries:
            last_row = len(industries) + 1
            industry_validation = DataValidation(
                type="list",
                formula1=f"=Industries!$A$2:$A${last_row}",
                allow_blank=True,
            )
            template_sheet.add_data_validation(industry_validation)
            industry_validation.add("D2:D500")

        nqf_last_row = len(nqf_levels) + 1
        nqf_validation = DataValidation(
            type="list",
            formula1=f"='NQF Levels'!$A$2:$A${nqf_last_row}",
            allow_blank=True,
        )
        template_sheet.add_data_validation(nqf_validation)
        nqf_validation.add("F2:F500")

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            "attachment; filename=occupation_template.xlsx"
        )
        return response

    if request.method == "POST":
        upload_file = request.FILES.get("file")
        if not upload_file:
            messages.error(request, "Please upload an Excel file.")
        elif not upload_file.name.endswith(".xlsx"):
            messages.error(request, "File must be an .xlsx Excel document.")
        else:
            try:
                from openpyxl import load_workbook

                workbook = load_workbook(upload_file, data_only=True)
                sheet = workbook.active

                count = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) < 2:
                        continue

                    ofo_code = str(row[0]).strip() if row[0] is not None else ""
                    ofo_title = str(row[1]).strip() if row[1] is not None else ""
                    description = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                    industry_code = (
                        str(row[3]).strip() if len(row) > 3 and row[3] else None
                    )
                    years_of_experience = 0
                    if len(row) > 4 and row[4] is not None:
                        try:
                            years_of_experience = int(row[4])
                        except (TypeError, ValueError):
                            years_of_experience = 0
                    preferred_nqf_level = 0
                    if len(row) > 5 and row[5] is not None:
                        try:
                            preferred_nqf_level = int(row[5])
                        except (TypeError, ValueError):
                            preferred_nqf_level = 0

                    if not ofo_code or not ofo_title:
                        continue

                    industry = None
                    if industry_code:
                        industry = Industry.objects.filter(code=industry_code).first()

                    Occupation.objects.update_or_create(
                        ofo_code=ofo_code,
                        defaults={
                            "ofo_title": ofo_title,
                            "description": description,
                            "industry": industry,
                            "years_of_experience": years_of_experience,
                            "preferred_nqf_level": preferred_nqf_level,
                        },
                    )
                    count += 1
                messages.success(request, f"Successfully uploaded {count} occupations.")
                return redirect("occupations")
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")

    context = navbar_context(request)
    return render(request, "content/occupation_upload.html", context)


@login_required
@user_passes_test(is_content_manager_or_admin)
def occupation_delete(request, occupation_id):
    """
    View to delete an occupation.
    """
    occupation = get_object_or_404(Occupation, pk=occupation_id)
    if request.method == "POST" or request.method == "DELETE":
        occupation.delete()
        messages.success(request, "Occupation deleted successfully.")
        return redirect("occupations")

    context = navbar_context(request)
    context["occupation"] = occupation
    return render(request, "content/occupation_confirm_delete.html", context)


@login_required
@user_passes_test(is_content_manager_or_admin)
def occupation_bulk_delete(request):
    """
    View to delete multiple occupations.
    """
    if request.method == "POST":
        occupation_ids = request.POST.getlist("selected_occupations")
        if occupation_ids:
            deleted_count = Occupation.objects.filter(id__in=occupation_ids).delete()[0]
            messages.success(
                request,
                f"Successfully deleted {deleted_count} occupations related records.",
            )
        else:
            messages.warning(request, "No occupations selected for deletion.")
    return redirect("occupations")


@login_required
def task_list_partial(request):
    """
    HTMX view to return a list of tasks for the search modal.
    """
    from django.db.models import Q
    from django.core.paginator import Paginator

    query = request.GET.get("q", "")
    tasks_qs = (
        OccupationTask.objects.values("id", "title", "description")
        .distinct()
        .order_by("title")
    )

    if query:
        tasks_qs = tasks_qs.filter(
            Q(title__icontains=query) | Q(description__icontains=query)
        )

    paginator = Paginator(tasks_qs, 10)
    page_number = request.GET.get("page")
    tasks_page = paginator.get_page(page_number)

    return render(
        request,
        "content/partials/task_list_selector.html",
        {"tasks": tasks_page, "search_query": query},
    )


@login_required
@user_passes_test(is_content_manager_or_admin)
def industry_list(request):
    query = request.GET.get("q", "")
    industries = Industry.objects.all().order_by("name")

    if query:
        industries = industries.filter(name__icontains=query)

    context = navbar_context(request)
    context.update({"industries": industries, "search_query": query})

    return render(request, "content/industry_list.html", context)


@login_required
@user_passes_test(is_content_manager_or_admin)
def industry_add(request):
    if request.method == "POST":
        form = IndustryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Industry created successfully!")
            return redirect("industry-list")
    else:
        form = IndustryForm()

    context = navbar_context(request)
    context["form"] = form
    return render(request, "content/industry_form.html", context)


@login_required
@user_passes_test(is_content_manager_or_admin)
def industry_edit(request, industry_id):
    industry = get_object_or_404(Industry, pk=industry_id)
    if request.method == "POST":
        form = IndustryForm(request.POST, instance=industry)
        if form.is_valid():
            form.save()
            messages.success(request, "Industry updated successfully!")
            return redirect("industry-list")
    else:
        form = IndustryForm(instance=industry)

    context = navbar_context(request)
    context.update({"form": form, "industry": industry})
    return render(request, "content/industry_form.html", context)


@login_required
@user_passes_test(is_content_manager_or_admin)
def industry_delete(request, industry_id):
    industry = get_object_or_404(Industry, pk=industry_id)
    if request.method == "POST" or request.method == "DELETE":
        industry.delete()
        messages.success(request, "Industry deleted successfully.")
        return redirect("industry-list")

    context = navbar_context(request)
    context["industry"] = industry
    return render(request, "content/industry_confirm_delete.html", context)


@login_required
@user_passes_test(is_content_manager_or_admin)
def skill_list(request):
    query = request.GET.get("q", "")
    skills = Skill.objects.all().order_by("name")

    if query:
        skills = skills.filter(name__icontains=query)

    context = navbar_context(request)
    context.update({"skills": skills, "search_query": query})
    return render(request, "content/skill_list.html", context)


@login_required
@user_passes_test(is_content_manager_or_admin)
def skill_add(request):
    if request.method == "POST":
        form = SkillForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Skill created successfully!")
            return redirect("skill-list")
    else:
        form = SkillForm()

    context = navbar_context(request)
    context["form"] = form
    return render(request, "content/skill_form.html", context)


@login_required
@user_passes_test(is_content_manager_or_admin)
def skill_edit(request, skill_id):
    skill = get_object_or_404(Skill, pk=skill_id)
    if request.method == "POST":
        form = SkillForm(request.POST, instance=skill)
        if form.is_valid():
            form.save()
            messages.success(request, "Skill updated successfully!")
            return redirect("skill-list")
    else:
        form = SkillForm(instance=skill)

    context = navbar_context(request)
    context.update({"form": form, "skill": skill})
    return render(request, "content/skill_form.html", context)


@login_required
@user_passes_test(is_content_manager_or_admin)
def skill_delete(request, skill_id):
    skill = get_object_or_404(Skill, pk=skill_id)
    if request.method == "POST" or request.method == "DELETE":
        skill.delete()
        messages.success(request, "Skill deleted successfully.")
        return redirect("skill-list")

    context = navbar_context(request)
    context["skill"] = skill
    return render(request, "content/skill_confirm_delete.html", context)
